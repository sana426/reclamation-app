# reclamations/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Count, Q, Avg
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from .models import Reclamation
from datetime import datetime, timedelta
import json
import openpyxl
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

# ==================== VUES PRINCIPALES ====================

@login_required
def accueil(request):
    return render(request, 'reclamations/accueil.html')

@login_required
def rechercher_reclamations(request):
    query = request.GET.get('q', '')
    region = request.GET.get('region', '')
    
    reclamations = Reclamation.objects.all()
    
    if query:
        reclamations = reclamations.filter(
            Q(id__icontains=query) |
            Q(nom__icontains=query) |
            Q(prenom__icontains=query) |
            Q(cin__icontains=query) |
            Q(telephone__icontains=query) |
            Q(numero_contrat__icontains=query)
        )
    
    if region and region != 'all':
        reclamations = reclamations.filter(agence=region)
    
    data = []
    for r in reclamations:
        data.append({
            'id': r.id,
            'nom': r.nom,
            'prenom': r.prenom or '',
            'cin': r.cin,
            'telephone': r.telephone,
            'region': r.agence,
            'type_probleme': r.get_type_probleme_display(),
            'urgence': r.get_urgence_display(),
            'statut': r.get_statut_display(),
            'statut_code': r.statut,
            'numero_contrat': r.numero_contrat or '',
            'date_creation': r.date_creation.strftime('%Y-%m-%d'),
        })
    
    return JsonResponse({'data': data, 'total': len(data)})

@login_required
def liste_reclamations(request):
    reclamations = Reclamation.objects.all()
    
    statut_filter = request.GET.get('statut', '')
    if statut_filter and statut_filter != 'all':
        reclamations = reclamations.filter(statut=statut_filter)
    
    agence_filter = request.GET.get('agence', '')
    if agence_filter and agence_filter != 'all':
        reclamations = reclamations.filter(agence=agence_filter)
    
    paginator = Paginator(reclamations, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'reclamations': page_obj,
        'total': reclamations.count(),
        'statut_filter': statut_filter,
        'agence_filter': agence_filter,
        'statut_choices': Reclamation.STATUT_CHOICES,
        'agence_choices': Reclamation.AGENCE_CHOICES,
    }
    return render(request, 'reclamations/liste.html', context)

@login_required
def changer_statut(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        reclamation_id = data.get('id')
        nouveau_statut = data.get('statut')
        
        reclamation = get_object_or_404(Reclamation, id=reclamation_id)
        reclamation.statut = nouveau_statut
        
        # Ajoute la date de clôture si le statut est "resolue"
        if nouveau_statut == 'resolue':
            reclamation.date_cloture = datetime.now().date()
        else:
            reclamation.date_cloture = None
        
        reclamation.save()
        
        return JsonResponse({'success': True})
    return JsonResponse({'success': False})

@login_required
@user_passes_test(lambda u: u.is_superuser)
def supprimer_reclamation(request, reclamation_id):
    reclamation = get_object_or_404(Reclamation, id=reclamation_id)
    reclamation.delete()
    messages.success(request, f'Réclamation {reclamation_id} supprimée')
    return redirect('reclamations:liste_reclamations')

@login_required
def detail_reclamation(request, reclamation_id):
    reclamation = get_object_or_404(Reclamation, id=reclamation_id)
    return render(request, 'reclamations/detail.html', {'reclamation': reclamation})

@login_required
def modifier_reclamation(request, reclamation_id):
    reclamation = get_object_or_404(Reclamation, id=reclamation_id)
    
    if request.method == 'POST':
        reclamation.nom = request.POST.get('nom', '')
        reclamation.prenom = request.POST.get('prenom', '')
        reclamation.cin = request.POST.get('cin', '')
        reclamation.telephone = request.POST.get('telephone', '')
        reclamation.email = request.POST.get('email', '')
        reclamation.agence = request.POST.get('agence', '')
        reclamation.type_probleme = request.POST.get('type_probleme', '')
        reclamation.urgence = request.POST.get('urgence', '')
        reclamation.adresse = request.POST.get('adresse', '')
        reclamation.numero_contrat = request.POST.get('numero_contrat', '')
        reclamation.description = request.POST.get('description', '')
        
        reclamation.save()
        messages.success(request, f'✅ Réclamation {reclamation_id} modifiée')
        return redirect('reclamations:liste_reclamations')
    
    nom_complet = reclamation.nom.split()
    prenom = nom_complet[-1] if len(nom_complet) > 1 else ''
    nom = nom_complet[0] if len(nom_complet) > 0 else ''
    
    context = {
        'reclamation': reclamation,
        'nom': nom,
        'prenom': prenom,
        'agences': Reclamation.AGENCE_CHOICES,
        'types': Reclamation.TYPE_PROBLEME,
        'urgences': Reclamation.URGENCE_CHOICES,
    }
    return render(request, 'reclamations/modifier.html', context)

@login_required
def nouvelle_reclamation(request):
    if request.method == 'POST':
        last_reclamation = Reclamation.objects.order_by('-id').first()
        if last_reclamation:
            last_num = int(last_reclamation.id.replace('REC-', ''))
            new_num = last_num + 1
        else:
            new_num = 1
        new_id = f'REC-{new_num:03d}'
        
        reclamation = Reclamation.objects.create(
            id=new_id,
            nom=request.POST.get('nom', ''),
            prenom=request.POST.get('prenom', ''),
            cin=request.POST.get('cin', ''),
            telephone=request.POST.get('telephone', ''),
            email=request.POST.get('email', ''),
            agence=request.POST['agence'],
            type_probleme=request.POST['type_probleme'],
            urgence=request.POST['urgence'],
            adresse=request.POST.get('adresse', ''),
            numero_contrat=request.POST.get('numero_contrat', ''),
            description=request.POST['description'],
            service=request.POST.get('service', 'eau'),
            cree_par=request.user
            
        )
        
        messages.success(request, f'✅ Réclamation {new_id} soumise')
        return redirect('reclamations:liste_reclamations')
    
    context = {
        'agences': Reclamation.AGENCE_CHOICES,
        'types': Reclamation.TYPE_PROBLEME,
        'urgences': Reclamation.URGENCE_CHOICES,
    }
    return render(request, 'reclamations/nouvelle.html', context)

@login_required
def dashboard(request):
    return render(request, 'reclamations/dashboard.html')

@login_required
def dashboard_stats(request):
    period = request.GET.get('period', 'all')
    agence_filter = request.GET.get('agence', 'all')
    
    reclamations = Reclamation.objects.all()
    
    now = datetime.now().date()
    if period == 'month':
        start_date = now - timedelta(days=30)
        reclamations = reclamations.filter(date_creation__gte=start_date)
    elif period == 'quarter':
        start_date = now - timedelta(days=90)
        reclamations = reclamations.filter(date_creation__gte=start_date)
    elif period == 'year':
        start_date = now - timedelta(days=365)
        reclamations = reclamations.filter(date_creation__gte=start_date)
    
    if agence_filter != 'all':
        reclamations = reclamations.filter(agence=agence_filter)
    
    total = reclamations.count()
    resolues = reclamations.filter(statut='resolue').count()
    taux_resolution = round((resolues / total * 100), 1) if total > 0 else 0
    temps_moyen_traitement = round(reclamations.filter(statut='resolue').count() * 2.5 / max(total, 1), 1)
    urgentes_non_resolues = reclamations.filter(statut__in=['en_attente', 'en_cours'], urgence='urgent').count()
    
    agence_stats = {}
    agences = ['Azilal', 'Afourer', 'Bzou', 'Demnate', 'Ait Attab', 'Tanante', 'Ouaouizeghth']
    for a in agences:
        count = reclamations.filter(agence=a).count()
        if count > 0:
            agence_stats[a] = count
    
    top_agence = {'nom': 'Aucune', 'count': 0}
    for a, count in agence_stats.items():
        if count > top_agence['count']:
            top_agence = {'nom': a, 'count': count}
    
    type_names = {
        'fuite': 'Fuite d\'eau', 'pression': 'Basse pression',
        'coupure': 'Coupure d\'eau', 'qualite': 'Qualité de l\'eau',
        'facturation': 'Facturation', 'autre': 'Autre'
    }
    type_stats = {}
    for t_code, t_name in type_names.items():
        count = reclamations.filter(type_probleme=t_code).count()
        if count > 0:
            type_stats[t_name] = count
    
    top_type = {'nom': 'Aucun', 'count': 0}
    for t, count in type_stats.items():
        if count > top_type['count']:
            top_type = {'nom': t, 'count': count}
    
    statut_stats = {
        'En attente': reclamations.filter(statut='en_attente').count(),
        'En cours': reclamations.filter(statut='en_cours').count(),
        'Clôturée': reclamations.filter(statut='cloturee').count()
    }
    
    trend_data = []
    for i in range(1, 13):
        count = reclamations.filter(date_creation__month=i).count()
        trend_data.append(count)
    
    avg_last_3 = sum(trend_data[-3:]) / 3 if len(trend_data) >= 3 else sum(trend_data) / max(len(trend_data), 1)
    prevision = round(avg_last_3 * 1.1)
    
    if len(trend_data) >= 4 and sum(trend_data[-2:]) > sum(trend_data[-4:-2]):
        diff = sum(trend_data[-2:]) - sum(trend_data[-4:-2])
        tendance = round((diff / sum(trend_data[-4:-2]) * 100), 1) if sum(trend_data[-4:-2]) > 0 else 0
    else:
        tendance = 0
    
    if total == 0:
        insights = "📭 Aucune donnée disponible"
        recommandations = "Ajoutez des réclamations"
    else:
        insights = f" {total} réclamations. ✅ Taux: {taux_resolution}%. 🔴 Urgentes: {urgentes_non_resolues}.  Prévision: {prevision}"
        if top_agence['count'] > 0:
            insights += f" {top_agence['nom']} est l'agence la plus critique."
        recommandations = f" Objectif: 85% (actuel: {taux_resolution}%).  Traiter {urgentes_non_resolues} urgences."
    
    return JsonResponse({
        'kpis': {
            'total': total,
            'taux_resolution': taux_resolution,
            'temps_moyen_traitement': temps_moyen_traitement,
            'urgentes_non_resolues': urgentes_non_resolues,
        },
        'agence_stats': agence_stats,
        'type_stats': type_stats,
        'statut_stats': statut_stats,
        'trend_data': trend_data,
        'prevision': prevision,
        'tendance': tendance,
        'top_agence': top_agence,
        'top_type': top_type,
        'insights': insights,
        'recommandations': recommandations,
    })

# ==================== EXPORTS SIMPLES ====================

def export_excel(request):
    reclamations = Reclamation.objects.all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Réclamations"
    
    # En-têtes
    ws['A1'] = 'N°'
    ws['B1'] = 'Nom'
    ws['C1'] = 'CIN'
    ws['D1'] = 'Téléphone'
    ws['E1'] = 'Agence'
    ws['F1'] = 'Type'
    ws['G1'] = 'Urgence'
    ws['H1'] = 'Date'
    ws['I1'] = 'Statut'
    
    # Données
    row = 2
    for r in reclamations:
        ws[f'A{row}'] = r.id
        ws[f'B{row}'] = r.nom
        ws[f'C{row}'] = r.cin or ''
        ws[f'D{row}'] = r.telephone or ''
        ws[f'E{row}'] = r.agence
        ws[f'F{row}'] = r.get_type_probleme_display()
        ws[f'G{row}'] = r.get_urgence_display()
        ws[f'H{row}'] = r.date_creation.strftime('%d/%m/%Y')
        ws[f'I{row}'] = r.get_statut_display()
        row += 1
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reclamations.xlsx"'
    wb.save(response)
    return response


def export_pdf(request):
    reclamations = Reclamation.objects.all()
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reclamations.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()
    
    # Titre
    elements.append(Paragraph("SRM Azilal - Liste des Réclamations", styles['Title']))
    elements.append(Spacer(1, 20))
    
    # Tableau
    data = [['N°', 'Nom', 'CIN', 'Tél', 'Agence', 'Date', 'Statut']]
    for r in reclamations:
        data.append([
            r.id, 
            r.nom, 
            r.cin or '-', 
            r.telephone or '-', 
            r.agence,
            r.date_creation.strftime('%d/%m/%Y'),
            r.get_statut_display()
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(table)
    doc.build(elements)
    return response